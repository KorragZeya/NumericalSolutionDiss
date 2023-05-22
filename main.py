import math
import numpy as np
import scipy as sp
import openpyxl

book = openpyxl.Workbook()
sheet = book.active

L1, L2, L3, L4 = 0.10, 0.20, 1.20, 1.00
L = L1 + L2 + L3 + L4
# L = 3
t = 1.4 * 86400
dt = 25
Nt = int(t / dt)
dx = 0.01
Nx = int(L / dx)
end_index = Nx - 1

"""
это функция радиационного солнечного потока. 
Если время соответствует ночи, то поток равен 0.
"""
def q(k, dt):
    if (k*dt//43200) % 2 == 0:
        qin = 700*np.sin((2 * 3.14 / 86400) * k * dt)
        if qin < 0:
            qin = 0
        return qin
    else:
        return 0


Tair = lambda k: 288 + 5 * np.sin((2 * 3.14 / 86400) * k * dt)


def q_rad(Tsoil, Tair, a, eps, k):
    return eps * 5.7 * 10 ** (-8)*(Tsoil[0]**4 - a*Tair(k)**4)


def q_conv(Tsoil, Tair, k):
    alpha = 10 * (Tsoil[0] - Tair(k)) ** 0.33 if Tsoil[0] > Tair(k) else 10 * (Tair(k) - Tsoil[0]) ** 0.33
    return alpha*(Tsoil[0] - Tair(k))


def calculate_one_time(K_h, M, X, C, *flows):
    global Nx, dx, dt, end_index
    for i in range(Nx):
        a = -(dt / dx ** 2) * K_h[i]
        c = -(dt / dx ** 2) * K_h[i + 1]
        b = C - c - a
        if i == 0:
            M[0][0] = b + a
            M[0][1] = c
        elif i == Nx - 1:
            M[end_index][end_index - 1] = a
            M[end_index][end_index] = b + c
        else:
            M[i][i - 1] = a
            M[i][i] = b
            M[i][i + 1] = c
    D = np.array([C * X[0] - a * (dx / K_h[0]) * sum(flows)
                  if i == 0 else C * X[i] for i in range(Nx)], dtype=float)
    X = np.copy(sp.sparse.csr_matrix(sp.linalg.inv(M)).dot(D))
    print(X[0])


def numerical_solution():
# Для уравнения теплопроводности
    Tsoil = np.full(Nx, 273)
    # lambdas_h = np.full(Nx + 1, 0.7)
    K1, K2, K3, K4 = 0.7, 0.7, 0.7, 0.7
    lambda_i = [0.07] + [K1 for _ in range(int(L1 / dx))] + \
               [K2 for _ in range(int(L2 /dx))] + \
               [K3 for _ in range(int(L3 / dx))] + \
               [K4 for _ in range(int(L4 / dx))] + [K4]

    lambda_h = np.array([2 * lambda_i[i] * lambda_i[i + 1] / (lambda_i[i] + lambda_i[i + 1]) for i in range(Nx + 1)])

    Cs = 2006200
    Cw = 4200000
    F = 0

    TMatrix = np.zeros((Nx, Nx), dtype=float)

    for k in range(Nt):
        calculate_one_time(lambda_h, TMatrix, Tsoil, Cs,
                           q(k, dt), q_rad(Tsoil, Tair, 0.8, 0.83, k), q_conv(Tsoil, Tair, k))

    for i in range(Nx):
        sheet.cell(row=i + 1, column=1).value = float(Tsoil[i]) - 273
        sheet.cell(row=i + 1, column=2).value = L - i * dx

    book.save("Numerical_solution.xlsx")
    book.close()


numerical_solution()
